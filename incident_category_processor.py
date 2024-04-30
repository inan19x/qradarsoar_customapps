import resilient
import re
import json 
import logging 
import smtplib
import ssl
import openpyxl
import os
import resilient_lib
import shutil
import email
import email.header
import email.mime.multipart

from datetime import datetime
from xml.dom import minidom

from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from circuits.core.handlers import handler
from resilient_circuits.actions_component import ResilientComponent, ActionMessage

logger = logging.getLogger(__name__)

class IncidentCategoryProcessor(ResilientComponent):
	# Subscribe to the Action Module message destination named 'incident_category_checker'
	channel = 'actions.incident_category_checker'

	@handler('incident_category_checker')
	def _incident_category_checker_handler_function(self, event, headers, *args, **kwargs): 
		
		incident = event.message['incident']
		incident_id = incident['id']
		incident_name = incident['name'].encode(encoding="utf-8",errors="strict").strip()

		logger.info('Category Checker - Incident {0}: {1}'.format(incident_id, incident_name))

		parser = resilient.ArgumentParser(config_file=resilient.get_config_file())
		opts = parser.parse_args()
		client = resilient.get_client(opts)

		path = '/home/resadmin/components/mapping_files/'
		filename = 'SOC_Playbook_Escalation.xlsx'

		mapper = self.get_incident_category_list(path, filename)
		
		if incident_name in mapper:
			client.context_header = headers['Co3ContextToken']
			uri = '/incidents/{0}'.format(incident_id)	

			incident = client.get(uri)

			# Report
			incident['properties']['incident_category'] = mapper[incident_name]
			
			logger.info('Updating Category <{0}> to be {1}'.format(incident_name, mapper[incident_name]))

			client.put(uri, incident)

	def create_dict_from_2_columns(self, workbook, worksheet, key_col, value_col):
		mapping = {}

		sheet_obj = workbook[worksheet]
		m_row = sheet_obj.max_row 
		
		keys = []
		values = []

		for i in range(2, m_row + 1): 
			cell_obj = sheet_obj.cell(row = i, column = key_col) 
			keys.append(cell_obj.value) 

		for i in range(2, m_row + 1): 
			cell_obj = sheet_obj.cell(row = i, column = value_col) 
			values.append(cell_obj.value) 

		counter = len(keys)
		count = 0

		while count < counter:
			mapping[keys[count]] = values[count]
			count += 1

		return mapping

	def get_incident_category_list(self, path, filename):
		workbook = openpyxl.load_workbook(path + filename)
		worksheet = 'Playbook'
		mapping_incident_vs_category = self.create_dict_from_2_columns(workbook, worksheet, 1, 2)

		return mapping_incident_vs_category