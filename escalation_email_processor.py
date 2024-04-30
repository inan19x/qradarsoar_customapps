import resilient
import re
import json 
import logging 
import smtplib
import ssl
import openpyxl
import os
import resilient_lib
import time
import shutil
import email
import email.header
import email.mime.multipart

from datetime import datetime
from xml.dom import minidom

from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from circuits.core.handlers import handler
from resilient_circuits.actions_component import ResilientComponent, ActionMessage

logger = logging.getLogger(__name__)

class EscalationEmailProcessor(ResilientComponent):
	# Subscribe to the Action Module message destination named 'escalation_email'
	channel = 'actions.escalation_email'


	# first escalation
	@handler('send_escalation_email')
	def _send_escalation_email_handler_function(self, event, headers, *args, **kwargs): 
		
		incident_detail	= self.get_incident_detail(event)

		logger.info("Sending Escalation Email from Incident ID {0} : {1}.".format(incident_detail['incident_id'], incident_detail['incident_name']))

		client = self.get_resilient_api_client(headers['Co3ContextToken'])
		config = self.get_config()
		
		incident_artifacts = self.get_incident_artifacts_processed(client, incident_detail['incident_id'])
		incident_attachments = self.get_incident_attachments(client, incident_detail['incident_id'])
		incident_artifacts_raws = self.get_incident_artifacts(client, incident_detail['incident_id'])

		mapper = self.get_resolver_group_detail(config['escalation_list_filepath'])
		email_info = self.get_email_info(mapper, incident_detail['resolver_team'])
		email_info = self.check_if_special_case_incident(config, email_info, incident_detail['offense_source'], incident_detail['resolver_team'], incident_artifacts_raws)
		email_data = self.process_email('first_escalation', config, email_info, incident_detail, incident_artifacts, incident_attachments)
		
		# Report
		while True:
			try:
				incident_data = self.get_incident_in_resi_by_id(client, incident_detail['incident_id'])		
				incident_data['properties']['successful_escalation'] = '; '.join(email_data['receiver_email'])
				incident_data['properties']['escalation_email_cc'] = '; '.join(email_data['cc_email'])

				if incident_data['properties']['failed_escalation'] is not None:
					incident_data['properties']['failed_escalation'] = '; '.join(list(set(incident_data['properties']['failed_escalation'].split('; ')) - set(email_data['receiver_email'])))
				
				if incident_data['properties']['resolver_team'] is not None:
					incident_data['properties']['resolver_team'] = incident_data['properties']['resolver_team']  + '; ' + incident_detail['resolver_team']
				else:
					incident_data['properties']['resolver_team'] = incident_detail['resolver_team']

				self.update_incident_in_resi_by_id(client, incident_detail['incident_id'], incident_data)
				break
			except:
				logger.info("Error updating incident data. Will try.")
				pass


	# reminder incident
	@handler('email_reminder_to_resolver_team')
	def _email_reminder_to_resolver_team_handler_function(self, event, headers, *args, **kwargs): 
		# gather basic info
		incident_detail	= self.get_incident_detail(event)

		logger.info("Sending Email Reminder for Incident ID {0} : {1}.".format(incident_detail['incident_id'], incident_detail['incident_name']))

		client = self.get_resilient_api_client(headers['Co3ContextToken'])
		config = self.get_config()
		
		# gather detail data
		incident_artifacts = self.get_incident_artifacts_processed(client, incident_detail['incident_id'])
		incident_attachments = self.get_incident_attachments(client, incident_detail['incident_id'])
		incident_artifacts_raws = self.get_incident_artifacts(client, incident_detail['incident_id'])
		
		# prepare and process email 
		mapper = self.get_resolver_group_detail(config['escalation_list_filepath'])
		
		email_info = {}
		to_addresses = []
		cc_addresses = []
		contents = []

		resolver_teams = incident_detail['resolver_id'].split('; ')
		for resolver_team in resolver_teams:
			info = self.get_email_info(mapper, resolver_team.strip())
			
			to_addresses += info['to_addresses']
			cc_addresses += info['cc_addresses']
			contents.append(info['content'])

		if len(contents) > 1:
			email_info['content'] = '<br>'.join(contents)
		else:
			email_info['content'] = ''.join(contents)

		# email_info['content'] = contents[0] + str(len(contents))

		email_info['to_addresses'] = to_addresses
		email_info['cc_addresses'] = cc_addresses

		email_info = self.check_if_special_case_incident(config, email_info, incident_detail['offense_source'], incident_detail['resolver_id'], incident_artifacts_raws)
		email_data = self.process_email('reminder', config, email_info, incident_detail, incident_artifacts, incident_attachments)

		# Report
		while True:
			try:
				incident_data = self.get_incident_in_resi_by_id(client, incident_detail['incident_id'])
				incident_data['properties']['email_reminder_count'] = email_data['notes']
				self.update_incident_in_resi_by_id(client, incident_detail['incident_id'], incident_data)
				break
			except:
				logger.info("Error updating incident data. Will try.")
				pass


	# closing incident
	@handler('send_notification_email')
	def _send_notification_email_handler_function(self, event, headers, *args, **kwargs): 
		# gather basic info
		incident_detail	= self.get_incident_detail(event)

		logger.info("Sending Closing Email for Incident ID {0} : {1}.".format(incident_detail['incident_id'], incident_detail['incident_name']))

		client = self.get_resilient_api_client(headers['Co3ContextToken'])
		config = self.get_config()
		
		# gather detail data
		incident_artifacts = self.get_incident_artifacts_processed(client, incident_detail['incident_id'])
		incident_attachments = self.get_incident_attachments(client, incident_detail['incident_id'])
		incident_artifacts_raws = self.get_incident_artifacts(client, incident_detail['incident_id'])
		
		# prepare and process email 
		mapper = self.get_resolver_group_detail(config['escalation_list_filepath'])
		
		email_info = {}
		to_addresses = []
		cc_addresses = []
		contents = []

		resolver_teams = incident_detail['resolver_id'].split('; ')
		for resolver_team in resolver_teams:
			info = self.get_email_info(mapper, resolver_team.strip())
			
			to_addresses += info['to_addresses']
			cc_addresses += info['cc_addresses']
			contents.append(info['content'])

		if len(contents) > 1:
			email_info['content'] = '<br>'.join(contents)
		else:
			email_info['content'] = ''.join(contents)

		# email_info['content'] = contents[0] + str(len(contents))

		email_info['to_addresses'] = to_addresses
		email_info['cc_addresses'] = cc_addresses

		email_info = self.check_if_special_case_incident(config, email_info, incident_detail['offense_source'], incident_detail['resolver_id'], incident_artifacts_raws)
		
		self.process_email('closing', config, email_info, incident_detail, incident_artifacts, incident_attachments)
		
		# Report
		while True:
			try:
				incident_data = self.get_incident_in_resi_by_id(client, incident_detail['incident_id'])		
				incident_data['properties']['notify_closed_incident_status'] = 'Yes'
				self.update_incident_in_resi_by_id(client, incident_detail['incident_id'], incident_data)
				break
			except:
				logger.info("Error updating incident data. Will try.")
				pass


	def get_config(self):
		config = {} # initiate var

		parser = resilient.ArgumentParser(config_file=resilient.get_config_file())

		# os.path.join(dir_name, base_filename)
		
		# email config
		config['smtp_server'] = parser.getopt('fn_outbound_email', 'smtp_server').encode('utf-8')
		config['smtp_port'] = parser.getopt('fn_outbound_email', 'smtp_port').encode('utf-8')
		config['smtp_email'] = parser.getopt('fn_outbound_email', 'smtp_email').encode('utf-8')
		config['smtp_user'] = parser.getopt('fn_outbound_email', 'smtp_user').encode('utf-8')
		config['smtp_password'] = parser.getopt('fn_outbound_email', 'smtp_password').encode('utf-8')
		config['message'] = email.mime.multipart.MIMEMultipart()

		# mapping files
		config['mail_structure_template_filepath'] = parser.getopt('resilient', 'mail_structure_template_filepath').encode('utf-8')
		config['mail_reminder_structure_template_filepath'] = parser.getopt('resilient', 'mail_reminder_structure_template_filepath').encode('utf-8')
		config['mail_close_incident_template_filepath'] = parser.getopt('resilient', 'mail_close_incident_template_filepath').encode('utf-8')
		config['escalation_list_filepath'] = parser.getopt('resilient', 'escalation_list_filepath').encode('utf-8')

		return config


	def get_resilient_api_client(self, Co3ContextToken):
		parser = resilient.ArgumentParser(config_file=resilient.get_config_file())
		opts = parser.parse_args()
		res_api_client = resilient.get_client(opts)

		res_api_client.context_header = Co3ContextToken

		return res_api_client


	def get_incident_in_resi_by_id(self, resi_client, incident_id):
		uri = '/incidents/{0}'.format(incident_id)
		incident = resi_client.get(uri)

		return incident


	def get_incident_artifacts(self, resi_client, incident_id):
		uri = "/incidents/{}/artifacts".format(incident_id)
		artifacts = resi_client.get(uri)

		return artifacts


	def get_incident_attachments(self, resi_client, incident_id):
		## Input attachments
		uri = '/incidents/{}/attachments'.format(incident_id)
		attachments = resi_client.get(uri)

		return attachments


	def update_incident_in_resi_by_id(self, resi_client, incident_id, incident_data):
		uri = '/incidents/{0}'.format(incident_id)
		resi_client.put(uri, incident_data)
		

	def get_incident_artifacts_processed(self, resi_client, incident_id):
		incident_artifacts = self.get_incident_artifacts(resi_client, incident_id)
		artifact_table = self.get_artifact_table()
		artifact_rows_template = self.get_artifact_rows_template()
		artifact_rows = ''
		artifacts_info = ''

		for artifact in incident_artifacts:
			if artifact['relating']:
				uri = '/artifact_types/{}'.format(artifact['type'])
				type_info = resi_client.get(uri)
				type_artifact = type_info['name']

				if artifact['type'] == 1 and artifact['properties'] is not None:
					try:
						type_artifact = type_info['name'] + ':' + artifact['properties'][0]['name']
					finally:
						type_artifact = type_info['name']

				value = artifact['value'].encode(encoding="utf-8",errors="strict").strip()
				description_ = artifact['description'].encode(encoding="utf-8",errors="strict").strip() if artifact['description'] is not None else ''
				artifact_rows = artifact_rows + artifact_rows_template.format(description_, value, type_artifact)

		if artifact_rows != '':
			artifacts_info = artifact_table.format(artifact_rows)

		return artifacts_info


	def get_email_info(self, mapper, resolver_team):
		# to address
		to_addresses = []
		cc_addresses = []
		content = []

		temp_email = mapper['to'][resolver_team] 
		temp_email = re.findall(r'[\w\.-]+@[\w\.-]+', temp_email)

		for email_addr in temp_email:
			to_addresses.append(email_addr.encode(encoding="utf-8",errors="strict").strip())

		# cc address
		temp_email = mapper['cc'][resolver_team] 
		temp_email = re.findall(r'[\w\.-]+@[\w\.-]+', temp_email)

		for email_addr in temp_email:
			cc_addresses.append(email_addr.encode(encoding="utf-8",errors="strict").strip())

		# content
		content = mapper['content'][resolver_team]

		return {'to_addresses':to_addresses, 'cc_addresses':cc_addresses, 'content':content}


	def get_incident_detail(self, event):
		incident_detail = {}
		incident = event.message['incident']
		incident_detail['incident_id'] = incident['id']

		# fill up QRadar SIEM ID
		incident_detail['incident_qid']	= 'N/A'

		if incident['properties']['qradar_id']:
			incident_detail['incident_qid']	= incident['properties']['qradar_id']

		incident_detail['incident_name'] = incident['name'].encode(encoding="utf-8",errors="strict").strip()
		incident_detail['resolver_id'] = incident['properties']['resolver_team']
		incident_detail['references'] = incident['properties']['references']
		incident_detail['successful_escalation']= incident['properties']['successful_escalation']
		incident_detail['escalation_email_cc'] = incident['properties']['escalation_email_cc']
		incident_detail['email_reminder_count'] = incident['properties']['email_reminder_count']
		incident_detail['successful_escalation']= incident['properties']['successful_escalation']
		incident_detail['escalation_email_cc'] = incident['properties']['escalation_email_cc']
		incident_detail['date_occured'] = datetime.fromtimestamp(float(incident['discovered_date'])/1000).strftime('%e %b %Y, %H:%M')
		incident_detail['description'] = incident['description'].encode(encoding="utf-8",errors="strict").strip()

		incident_detail['offense_source'] = ''
		if incident['properties']['offense_source'] is not None:
			incident_detail['offense_source'] = incident['properties']['offense_source'].encode(encoding="utf-8",errors="strict").strip()

		incident_detail['date_occured'] = datetime.fromtimestamp(float(incident['discovered_date'])/1000).strftime('%e %b %Y, %H:%M')
		
		incident_detail['notes'] = ''
		incident_detail['resolver_team'] = ''
		incident_detail['resolution_summary'] = ''

		if event.message['properties'] is not None:
			if 'notes' in event.message['properties']:
				incident_detail['notes'] = event.message['properties']['notes']

			if 'resolver_team' in event.message['properties']:
				resolver_id = event.message['properties']['resolver_team']
				incident_detail['resolver_team'] = event.message['type_info']['actioninvocation']['fields']['resolver_team']['values'][str(resolver_id)]['label']	

			if 'resolution_summary' in event.message['properties']:
				incident_detail['resolution_summary'] = event.message['properties']['resolution_summary']

		return incident_detail


	def process_email(self, type, config, email_info, incident_detail, incident_artifacts='<p>None</p>', incident_attachments=None):
		data = {}
		
		incident_id = incident_detail['incident_id']
		incident_qid = incident_detail['incident_qid']
		incident_name = incident_detail['incident_name']
		offense_source = incident_detail['offense_source']
		resolver_team = incident_detail['resolver_team']
		references = incident_detail['references']
		email_reminder_count= incident_detail['email_reminder_count']
		description = incident_detail['description']
		date_occured = incident_detail['date_occured']
		notes = incident_detail['notes']
		artifacts = incident_artifacts

		to_addresses = []
		cc_addresses = []
		content = []

		for email_addr in email_info['to_addresses']:
			to_addresses.append(email_addr.encode(encoding="utf-8",errors="strict").strip())

		for email_addr in email_info['cc_addresses']:
			cc_addresses.append(email_addr.encode(encoding="utf-8",errors="strict").strip())

		# content
		content = email_info['content']

		content_data = {'to': to_addresses, 'cc': cc_addresses, 'task': content}
		task = content_data['task']

		smtp_server = config['smtp_server']
		port = config['smtp_port']
		sender_email = config['smtp_email']
		receiver_email = content_data['to']
		cc_email = content_data['cc']
		all_receivers = receiver_email + cc_email
		username = config['smtp_user']
		password = config['smtp_password']
		message = config['message']

		message.set_charset('utf-8')
		
		# Create a secure SSL context
		context = ssl.create_default_context()

		recommendation_header = '<h3>Recommendation</h3>'

		filename = None

		subject_tag = ''

		if type == 'first_escalation':
			filename = config['mail_structure_template_filepath'] 
		elif type == 'reminder':
			filename = config['mail_reminder_structure_template_filepath']
			reminder_count = 1
			if not email_reminder_count:
				reminder_count = reminder_count
			else:
				reminder_count = int(email_reminder_count) + 1

			notes = reminder_count

			subject_tag = '[Reminder]'

		elif type == 'closing':
			filename = config['mail_close_incident_template_filepath']
			subject_tag = '[Closing]'

		elif type == 'confirmation':
			subject_tag = '[Confirmation]'

		elif type == 'info':
			filename = config['mail_structure_template_filepath'] 
			subject_tag = '[Info]'

		fo = open(filename, 'rb')
		filecontent = fo.read()

		message['Subject'] = email.header.Header('{0} Incident {1} - {2}'.format(subject_tag, incident_id, incident_name).strip())
		message['From'] = email.header.Header(sender_email)
		message['To'] = email.header.Header(', '.join(receiver_email).strip())
		message['Cc'] = email.header.Header(', '.join(cc_email).strip())
		message['Message-ID'] = email.header.Header(email.utils.make_msgid())
		message['Date'] = email.header.Header(email.utils.formatdate())
		
		css_style = self.get_css_style()

		if type == 'closing':
			text = filecontent.format(css_style, incident_id, incident_qid, incident_name, date_occured, incident_detail['resolution_summary'])
		else:	
			text = filecontent.format(css_style, recommendation_header, task, incident_id, incident_qid, 
						references, incident_name, offense_source, description, date_occured, notes, artifacts)

		part1 = MIMEText(text, 'html')
		
		message.attach(part1)
		
		for attachment in incident_attachments:
			# Get the file
			attachment_input = resilient_lib.get_file_attachment(incident_id=incident_id, 
				attachment_id=attachment['id'], res_client=self.rest_client())

			# Get the filename
			attachment_name = resilient_lib.get_file_attachment_name(incident_id=incident_id, 
				attachment_id=attachment['id'], res_client=self.rest_client())

			try:
				part2 = MIMEBase('application', 'octet-stream')
				part2.set_payload(attachment_input)
				encoders.encode_base64(part2)
				part2.add_header(
				    'Content-Disposition',
				    'attachment; filename= {}'.format(attachment_name)
				)
				message.attach(part2)
			finally:
				logger.debug('Adding attachment files done')

		# Try to log in to server and send email
		server = smtplib.SMTP(smtp_server,port)
		server.starttls() # Secure the connection
		server.login(username, password)

		# TODO: Send email here
		server.sendmail(sender_email, all_receivers, message.as_string())
		server.quit()

		# populate data for result
		data['receiver_email'] = receiver_email
		data['cc_email'] = cc_email
		data['notes'] = notes

		return data 


	def create_dict_from_2_columns(self, workbook, worksheet, key_col, value_col):
		mapping = {}

		sheet_obj = workbook[worksheet]
		m_row = sheet_obj.max_row 
		
		keys = []
		values = []

		for i in range(2, m_row + 1): 
			cell_obj = sheet_obj.cell(row = i, column = key_col) 
			keys.append(cell_obj.value) 

		for i in range(2, m_row + 1): 
			cell_obj = sheet_obj.cell(row = i, column = value_col) 
			values.append(cell_obj.value) 

		counter = len(keys)
		count = 0

		while count < counter:
			mapping[keys[count]] = values[count]
			count += 1

		return mapping


	def get_resolver_group_detail(self, file_path):
		workbook = openpyxl.load_workbook(file_path)
		worksheet = 'Email Template'
		mapping_resolver_vs_email_to = self.create_dict_from_2_columns(workbook, worksheet, 1, 2)
		mapping_resolver_vs_email_cc = self.create_dict_from_2_columns(workbook, worksheet, 1, 3)
		mapping_resolver_vs_email_content = self.create_dict_from_2_columns(workbook, worksheet, 1, 4)

		return {'to': mapping_resolver_vs_email_to, 'cc': mapping_resolver_vs_email_cc, 'content': mapping_resolver_vs_email_content}


	def get_css_style(self):
		css_style = """
			body {
				font-family: calibri, arial;
			}
			table, th, td {
				border: 1px solid black;
				border-collapse: collapse;
				font-size: 12pt;
			}
			th {
				background-color: black;
				color: #ffffff;
				text-align: center;
			}
			th, td {
			  	padding: 2px;
			}
			h2 {
				font-size: 18pt;
			}
			h3 {
				font-size: 14pt;
			}
			p { 
				margin: 1px;
				font-size: 12pt;
			}
		"""

		return css_style


	def get_artifact_table(self):

		artifact_table = """
			<table style='width:100%;border: 1px solid black;border-collapse: collapse;'>
				<tr>
					<th>Description</th>
					<th>Value</th>
					<th>Type</th>
				</tr>
				{0}
			</table>
			"""
		return artifact_table


	def get_artifact_rows_template(self):
		
		artifact_rows_template = """
				<tr>
					<td>{0}</td>
					<td>{1}</td>
					<td>{2}</td>
				</tr>
				"""
		return artifact_rows_template